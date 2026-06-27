import csv
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from data.export import build_payment_records, export_to_csv, export_to_excel
from domain.entities import Installment, Payment


def make_installment_with_payments():
    return Installment(
        loan_id=1,
        number=1,
        due_date=date(2026, 1, 8),
        scheduled_amount=Decimal("1000.00"),
        payments=[
            Payment(
                installment_id=None,
                payment_date=date(2026, 1, 8),
                amount_paid=Decimal("400.00"),
                method="Efectivo",
                notes="Abono parcial",
            ),
            Payment(
                installment_id=None,
                payment_date=date(2026, 1, 15),
                amount_paid=Decimal("600.00"),
                method="Transferencia",
            ),
        ],
    )


def test_build_payment_records_sorts_most_recent_first():
    installment = make_installment_with_payments()

    records = build_payment_records([installment], {1: "Juan Perez"})

    assert len(records) == 2
    assert records[0].payment_date == date(2026, 1, 15)
    assert records[0].amount_paid == Decimal("600.00")
    assert records[0].borrower_name == "Juan Perez"
    assert records[1].payment_date == date(2026, 1, 8)


def test_export_to_csv_writes_header_and_rows(tmp_path):
    installment = make_installment_with_payments()
    records = build_payment_records([installment], {1: "Juan Perez"})
    path = tmp_path / "historial.csv"

    export_to_csv(records, str(path))

    with open(path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    assert rows[0][0] == "Deudor"
    assert len(rows) == 3
    assert rows[1][0] == "Juan Perez"
    assert rows[1][6] == "600.00"


def test_export_to_excel_writes_header_and_rows(tmp_path):
    installment = make_installment_with_payments()
    records = build_payment_records([installment], {1: "Juan Perez"})
    path = tmp_path / "historial.xlsx"

    export_to_excel(records, str(path))

    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][0] == "Deudor"
    assert len(rows) == 3
    assert rows[1][0] == "Juan Perez"
    assert rows[1][6] == "600.00"
